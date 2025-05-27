from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from pyexpat.errors import messages
from django.contrib import messages


from .models import *
from openpyxl import load_workbook
# Create your views here.

from django.contrib.auth.decorators import login_required

@login_required(login_url='login')
def dashboard(request):
    # Sizning dashboard kodi
    return render(request, 'dashboard/dashboard.html')


def index(request):
    return render(request, 'index.html')

def form(request):
    y = Yunalish.objects.all()
    k = Kursi.objects.filter()
    g = Guruh.objects.all()
    return render(request, 'form.html', {
        'yunalishlar': y,
        'kurslar': k,
        'guruhlar': g,
    })

def get_kurslar(request):
    yunalish_id = request.GET.get('yunalish_id')
    kurslar = Kursi.objects.filter(guruh__yunalish__id=yunalish_id).distinct()
    data = [{'id': k.id, 'name': k.name} for k in kurslar]
    return JsonResponse({'kurslar': data})

def get_guruhlar(request):
    yunalish_id = request.GET.get('yunalish_id')
    kurs_id = request.GET.get('kurs_id')
    guruhlar = Guruh.objects.filter(
        yunalish__id=yunalish_id,
        kursi__id=kurs_id
    ).distinct()
    data = [{'id': g.id, 'name': g.name} for g in guruhlar]
    return JsonResponse({'guruhlar': data})

def get_fanlar(request):
    guruh_id = request.GET.get('guruh_id')
    fanlar = Fan.objects.filter(guruh__id=guruh_id).distinct()
    data = [{'id': f.id, 'name': f.name} for f in fanlar]
    return JsonResponse({'fanlar': data})

# def saqlash(request):
#     return render(request, 'test.html')


from random import sample
from .models import Test


def get_random_questions(fan_id):
    # Ma'lum fan bo'yicha 25 ta tasodifiy savol olish
    questions = Test.objects.filter(fan_id=fan_id)

    # Agar savollar soni 25 dan ko'proq bo'lsa, tasodifiy tanlash
    if questions.count() > 25:
        questions = sample(list(questions), 25)

    return questions

def start_test(request):
    if request.method == 'POST':
        fan_id = request.POST.get('fan_id')
        if not fan_id:
            return render(request, 'form.html', {
                'error': 'Fanni tanlang'
            })
        else:
            fan_id = int(fan_id)
            questions = get_random_questions(fan_id)
        return render(request, 'start_test.html', {
            'fan': fan_id,
            'questions': questions,
        })
    return render(request, 'form.html')



def test_views(request, id):
    fan = get_object_or_404(Fan, id=id)
    questions = Test.objects.filter(fan=fan)
    if questions.count() > 25:
        questions = sample(list(questions), 25)

    if request.method == "POST":
        name = request.POST.get('name')
        # fan= request.POST.get('fan_id')
        kurs = get_object_or_404(Kursi,id=request.POST.get('kurs_id'))
        guruh = get_object_or_404(Guruh, id=request.POST.get('guruh_id'))
        yonalish = get_object_or_404(Yunalish, id=request.POST.get('yunalish_id'))




    return render(request, 'tests.html', {
        'fan': fan,
        'name': name,
        'kurs': kurs,
        'guruh': guruh,
        'yonalish': yonalish,
        'questions': questions,  # fan_id bilan bir xil bo'lishi kerak emas, lekin xohlasangiz yuborish mumkin
    })



def submit_test(request):


    if request.method == "POST":
        id = request.POST.get('fan_id')
        fan = get_object_or_404(Fan, id=id)
        questions = Test.objects.filter(fan=fan)
        name = request.POST.get('name')
        fan= request.POST.get('fan_id')
        kurs = get_object_or_404(Kursi,id=request.POST.get('kurs_id'))
        guruh = get_object_or_404(Guruh, id=request.POST.get('guruh_id'))
        yonalish = get_object_or_404(Yunalish, id=request.POST.get('yunalish_id'))

        togrilar = 0

        for question in questions:
            tanlangan = request.POST.get(f'answer_{question.id}')
            if tanlangan == question.correct:
                togrilar += 1

        # Natijani saqlash
        Natija.objects.create(
            fan=fan,
            ism=name,
            kurs=kurs,
            guruh=guruh,
            yonalish=yonalish,
            togri_soni=togrilar,
            umumiy_savollar=questions.count()
        )
        return render(request, 'natija.html', {
            'ism': name,
            'familiya': name,
            'kurs': kurs,
            'guruh': guruh,
            'yonalish': yonalish,
            'togri': togrilar,
            'umumiy': questions.count(),
            'fan': fan,
        })


    else:
        return redirect('start_test', fan_id=id)




def natija(request):
    return render(request, 'natija.html')

def upload_test(request):
    fanlar = Fan.objects.all()

    if request.method == "POST" and request.FILES.get("excel_file"):
        fan_id = request.POST.get("fan_id")
        try:
            fan_obj = Fan.objects.get(id=fan_id)
        except Fan.DoesNotExist:
            messages.error(request, "Tanlangan fan topilmadi.")
            return redirect('upload_test')

        excel_file = request.FILES["excel_file"]
        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0] or not row[1]:
                    messages.warning(request, f"{idx}-qatorda savol yoki to‘g‘ri javob yo‘q — o‘tkazib yuborildi.")
                    continue

                test = Test.objects.create(
                    question=row[0],
                    correct=row[1],
                    wrong1=row[2] or '',
                    wrong2=row[3] or '',
                    wrong3=row[4] or '',
                )
                test.fan.set([fan_obj])

            messages.success(request, "Testlar saqlandi!")
            return redirect("upload_test")

        except Exception as e:
            messages.error(request, f"Xatolik yuz berdi: {e}")

    return render(request, "upload.html", {"fanlar": fanlar})





